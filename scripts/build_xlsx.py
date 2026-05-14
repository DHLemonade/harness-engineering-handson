#!/usr/bin/env python3
"""classification.json → output/classification_stats.xlsx

xlsx skill 규약을 따라:
- 시트별 책임 분리 (Summary / ByCategory / ByUrgency / ByChannel / HighUrgencyIds)
- 합계는 SUM 수식으로 (하드코딩 금지)
- 비율은 SUMIF/카운트 기반 수식으로
- BarChart(카테고리) + PieChart(긴급도) + BarChart(채널) 삽입
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

DATA = json.loads(Path("output/classification.json").read_text(encoding="utf-8"))

wb = Workbook()

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
BODY_FONT = Font(name="Arial", size=11)
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(cell) -> None:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER


# ---------- Sheet 1: Summary ----------
ws = wb.active
ws.title = "Summary"
ws["A1"] = "VoC 일괄 처리 통계"
ws["A1"].font = Font(name="Arial", size=14, bold=True)
ws.merge_cells("A1:C1")

ws["A3"] = "지표"
ws["B3"] = "값"
for c in ("A3", "B3"):
    style_header(ws[c])

ws["A4"] = "처리 건수(total)"
ws["B4"] = "=ByCategory!B10"  # 합계 셀 참조 (아래 정의)
ws["A5"] = "HIGH 긴급도 건수"
ws["B5"] = "=ByUrgency!B2"
ws["A6"] = "email 채널 건수"
ws["B6"] = "=ByChannel!B2"
ws["A7"] = "chat 채널 건수"
ws["B7"] = "=ByChannel!B3"

for r in range(4, 8):
    ws[f"A{r}"].font = BODY_FONT
    ws[f"B{r}"].font = BODY_FONT

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 14

# ---------- Sheet 2: ByCategory ----------
cat_ws = wb.create_sheet("ByCategory")
cat_ws["A1"] = "카테고리"
cat_ws["B1"] = "건수"
cat_ws["C1"] = "비율"
for c in ("A1", "B1", "C1"):
    style_header(cat_ws[c])

CATEGORIES = ["REFUND_REQUEST", "PAYMENT_INQUIRY", "SIGNUP_ISSUE",
              "DELIVERY_INQUIRY", "PRODUCT_INQUIRY", "COMPLAINT", "OTHER"]
for i, cat in enumerate(CATEGORIES, start=2):
    cat_ws.cell(row=i, column=1, value=cat).font = BODY_FONT
    cat_ws.cell(row=i, column=2, value=DATA["by_category"].get(cat, 0)).font = BODY_FONT
    # 비율 수식: 자기 건수 / 합계
    cat_ws.cell(row=i, column=3, value=f"=B{i}/$B$10").font = BODY_FONT
    cat_ws.cell(row=i, column=3).number_format = "0.0%"

# 합계 행
cat_ws["A10"] = "합계"
cat_ws["A10"].font = Font(name="Arial", size=11, bold=True)
cat_ws["B10"] = f"=SUM(B2:B{1 + len(CATEGORIES)})"
cat_ws["B10"].font = Font(name="Arial", size=11, bold=True)

cat_ws.column_dimensions["A"].width = 22
cat_ws.column_dimensions["B"].width = 10
cat_ws.column_dimensions["C"].width = 10

# 카테고리 BarChart
bar = BarChart()
bar.type = "bar"
bar.style = 11
bar.title = "카테고리별 건수"
bar.x_axis.title = "건수"
bar.y_axis.title = "카테고리"
data_ref = Reference(cat_ws, min_col=2, min_row=1, max_row=1 + len(CATEGORIES))
cats_ref = Reference(cat_ws, min_col=1, min_row=2, max_row=1 + len(CATEGORIES))
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.height = 9
bar.width = 16
cat_ws.add_chart(bar, "E1")

# ---------- Sheet 3: ByUrgency ----------
urg_ws = wb.create_sheet("ByUrgency")
urg_ws["A1"] = "긴급도"
urg_ws["B1"] = "건수"
for c in ("A1", "B1"):
    style_header(urg_ws[c])

URGENCIES = ["HIGH", "MEDIUM", "LOW"]
for i, u in enumerate(URGENCIES, start=2):
    urg_ws.cell(row=i, column=1, value=u).font = BODY_FONT
    urg_ws.cell(row=i, column=2, value=DATA["by_urgency"].get(u, 0)).font = BODY_FONT

urg_ws.column_dimensions["A"].width = 14
urg_ws.column_dimensions["B"].width = 10

pie = PieChart()
pie.title = "긴급도 분포"
data_ref = Reference(urg_ws, min_col=2, min_row=1, max_row=1 + len(URGENCIES))
labels_ref = Reference(urg_ws, min_col=1, min_row=2, max_row=1 + len(URGENCIES))
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(labels_ref)
pie.height = 9
pie.width = 12
urg_ws.add_chart(pie, "D1")

# ---------- Sheet 4: ByChannel ----------
ch_ws = wb.create_sheet("ByChannel")
ch_ws["A1"] = "채널"
ch_ws["B1"] = "건수"
for c in ("A1", "B1"):
    style_header(ch_ws[c])

CHANNELS = ["email", "chat"]
for i, ch in enumerate(CHANNELS, start=2):
    ch_ws.cell(row=i, column=1, value=ch).font = BODY_FONT
    ch_ws.cell(row=i, column=2, value=DATA["by_channel"].get(ch, 0)).font = BODY_FONT

ch_ws.column_dimensions["A"].width = 14
ch_ws.column_dimensions["B"].width = 10

bar2 = BarChart()
bar2.type = "col"
bar2.style = 12
bar2.title = "채널별 건수"
data_ref = Reference(ch_ws, min_col=2, min_row=1, max_row=1 + len(CHANNELS))
cats_ref = Reference(ch_ws, min_col=1, min_row=2, max_row=1 + len(CHANNELS))
bar2.add_data(data_ref, titles_from_data=True)
bar2.set_categories(cats_ref)
bar2.height = 8
bar2.width = 12
ch_ws.add_chart(bar2, "D1")

# ---------- Sheet 5: HighUrgencyIds ----------
hi_ws = wb.create_sheet("HighUrgencyIds")
hi_ws["A1"] = "voc_id"
hi_ws["B1"] = "처리 우선순위"
for c in ("A1", "B1"):
    style_header(hi_ws[c])

for i, vid in enumerate(DATA.get("high_urgency_ids", []), start=2):
    hi_ws.cell(row=i, column=1, value=vid).font = BODY_FONT
    hi_ws.cell(row=i, column=2, value="HIGH").font = BODY_FONT

hi_ws.column_dimensions["A"].width = 14
hi_ws.column_dimensions["B"].width = 16

# ---------- 저장 ----------
out_path = Path("output/classification_stats.xlsx")
wb.save(out_path)
print(f"생성: {out_path} (시트 {len(wb.sheetnames)}개: {wb.sheetnames})")
